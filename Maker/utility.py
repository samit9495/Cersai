import datetime
import errno
import re
import sys
import time
import urllib
import numpy
import os
import pickle
import pandas
import pyautogui
import collections
from shutil import copyfile
from selenium.webdriver.support import ui
from openpyxl import Workbook
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
import requests
from email.mime.text import MIMEText
import logging
from logging.handlers import RotatingFileHandler

# maker_id = "F010200059"

def page_is_loaded(driver):
    return driver.find_element_by_tag_name("body") != None


def submit_req(body, id1):
    tt = int(time.time() * 1000)
    # url = (F"http://htlms.spinup.tech/CERSAI/IBACRAddSIRecordServlet?token={tt}")
    url = (F"https://www.cersai.org.in/CERSAI/IBACRAddSIRecordServlet?token={tt}")
    headers = {
        "Connection": "keep-alive",
        "Accept": "text/html, */*; q=0.01",
        "Origin": "https://www.cersai.org.in",
        "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36",
        "Content-Type": "application/x-www-form-urlencoded",
        "Referer": "https://www.cersai.org.in/CERSAI/IBACRPageLoaderServlet",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "en-US,en;q=0.9",
        "Cookie": f"JSESSIONID={id1}",
    }
    # response = requests.post(url, data=body, headers=headers, verify=False)
    response = requests.post(url, data=body, headers=headers)
    print(response.text)
    st_code = response.status_code
    return response.text, st_code


def date_format(date):
    return str(datetime.datetime.date(date).strftime("%d-%m-%Y"))


def get_data(record,sid,maker_id):
    dd = datetime.datetime.now()-record[0][26]
    ndays = dd.days
    total_loan_amt = 0
    # for line in record:
    #     total_loan_amt += int(line[36])

    st1 = "BEACON_SCHEME_FLAG=Y&NATURE_OF_PROPERTY=0{}&BEACON_FLAG=G&SATN_FLAG=&SUBMITTED=false&MATCHING_ASSET_IDS=&SELECTED_ASSET_ID=0&ENTITY_LOGGED_IN_TYPE=HFC&PRIMARY_CHARGE_HOLDER_CODE={}&PRIMARY_CHARGE_HOLDER_PAN=AAFCM1099K&PRIMARY_CHARGE_HOLDER_EMAIL=mehra.anuj@mahindra.com&LINK_ID=&HELPER_HIDDEN=IBACRAddSIRecordHelper&SCREEN_ID=01501&SESSION_ID={}&LOGGED_IN_USER_ID={}&CONDONATION_FORM_SECURED_CREDITOR_NAME=MAHINDRA+RURAL+HOUSING+FINANCE+LTD&CONDONATION_FORM_SECURED_CREDITOR_ADDERSS=Branch+Name+:+{},++Building+Name+:+,+Street+Name/No+:+,+Village/Suburbs+:+,+Taluka+:+,+City+:+,+District+:+{},+PINCODE+:+{}&CONDONATION_FORM_BORROWER_NAME={}+and+others&CONDONATION_FORM_BORROWER_ADDRESS=Plot+No/Survey+No/Shop/Flat/House+No+:+{},++Building+No+:+,++Building+Name+:+,++Street+No./Name+:++,++Village+:+,++Locality+:+,++Taluka+:+,+District+:+{},+PINCODE+:+{}&CONDONATION_FORM_AMOUNT_OF_LOAN={}&CONDONATION_FORM_TRANSACTION_TYPE=Add+Security+Interest+&CONDONATION_FORM_DATE_OF_SI_CREATION={}&CONDONATION_FORM_NO_OF_DAYS_OF_DELAY={}&CONDONATION_FORM_REASON_OF_DELAY=Operational+Issue&CONDONATION_FORM_OMMISION_MISSING_STATEMENT=&CONDONATION_FORM_TRANSACTION_TYPE_ID=015&IBACRCondonationApplicationCheckerStatus=NO&PROCESS_TYPE=015&PROCESS_TYPE_DESC=Add+Security+Interest+&TOTAL_TABS=7&".format(
        str(record[0][1])[-1:], maker_id[:5], sid, maker_id, record[0][21], record[0][22], record[0][24], record[0][13],
        record[0][15], record[0][16], record[0][18], record[0][36],
        date_format(record[0][26]), ndays)

    st2 = ""
    for i,line in enumerate(record):
        if "individual" in str(line[10]).lower():
            btype = "IND"
        elif "hindu" in str(line[10]).lower():
            btype = "HUF"
        elif "company" in str(line[10]).lower():
            btype = "COM"
        elif "society" in str(line[10]).lower():
            btype = "COS"
        elif "liablity" in str(line[10]).lower():
            btype = "LLP"
        elif "partnership firm" in str(line[10]).lower():
            btype = "PAF"
        elif "proprietorship" in str(line[10]).lower():
            btype = "PRF"
        elif "trust" in str(line[10]).lower():
            btype = "TRS"

        title = str(line[12]).lower().replace(".", "")

        # plotno = urllib.urlencode(line[15])

        st2 += F"BORROWER_TYPE={btype}&BORROWER_PROP_OWNER={str(line[11]).upper()}&TIPVALUE_BORROWER_PROP_OWNER=&BORROWER_UID=&TIPVALUE_BORROWER_UID=&BORROWER_IND_TITLE_{i+1}={'Mr' if title == 'mr' else 'Mrs' if title == 'mrs' else 'Ms' if title == 'ms' else 'Dr'}.&TIPVALUE_BORROWER_IND_TITLE_{i+1}=&BORROWER_IND_SURNAME_{i+1}=&TIPVALUE_BORROWER_IND_SURNAME_{i+1}=&BORROWER_IND_NAME_{i+1}={line[13]}&TIPVALUE_BORROWER_IND_NAME_{i+1}=&BORROWER_IND_FATHER_HUSBAND_{i+1}=&TIPVALUE_BORROWER_IND_FATHER_HUSBAND_{i+1}=&BORROWER_IND_DOB_{i+1}={date_format(line[14])}&TIPVALUE_BORROWER_IND_DOB_{i+1}=&BORROWER_PLOTNO={line[15]}&TIPVALUE_BORROWER_PLOTNO=&BORROWER_BUILDINGNO=&TIPVALUE_BORROWER_BUILDINGNO=&BORROWER_BUILDINGNAME=&TIPVALUE_BORROWER_BUILDINGNAME=&BORROWER_STREET=&TIPVALUE_BORROWER_STREET=&BORROWER_VILLAGE=&TIPVALUE_BORROWER_VILLAGE=&BORROWER_LOCALITY=&TIPVALUE_BORROWER_LOCALITY=&BORROWER_TALUKA=&TIPVALUE_BORROWER_TALUKA=&BORROWER_DISTRICT={line[16]}&TIPVALUE_BORROWER_DISTRICT=&BORROWER_STATE={line[17] if int(line[17])>9 else F'0{line[17]}'}&TIPVALUE_BORROWER_STATE=&BORROWER_PINCODE={line[18]}&TIPVALUE_BORROWER_PINCODE=&BORROWER_PHONE=&TIPVALUE_BORROWER_PHONE=&BORROWER_PAN={line[19]}&TIPVALUE_BORROWER_PAN=&"

    unit = "Square+feet" if "feet" in str(record[0][4]).lower() or "ft" in str(
        record[0][4]).lower() else "Square+metre" if "metre" in str(record[0][4]).lower() else "Acre" if "acre" in str(
        record[0][4]).lower() else "Hectare" if "hectare" in str(record[0][4]).lower() else "Gunta" if "gunta" in str(
        record[0][4]).lower() else "Ares" if "ares" in str(record[0][4]).lower() else "Cents" if "cents" in str(
        record[0][4]).lower() else "Square+yard" if "yard" in str(record[0][4]).lower() else ""


    st3 = F"TPM_DELETION_FLAG=false&ED_FLAG=false&TPM_FLAG_HIDDEN=TPM_NO&JOINT_CHARGE=N&CONSORTIUM_FINANCE=N&NO_OF_CHARGE_HOLDERS=1&TIPVALUE_NO_OF_CHARGE_HOLDERS=&TIPVALUE_JC_CF=&JC_CF=&TIPVALUE_JOINT_CHARGE_YES=&TIPVALUE_JOINT_CHARGE_NO=&TIPVALUE_CONSORTIUM_FINANCE_YES=&TIPVALUE_CONSORTIUM_FINANCE_NO=&CHARGE_HOLDER=MAHINDRA+RURAL+HOUSING+FINANCE+LTD&CHARGE_HOLDER_BRANCH_CODE={record[0][20]}&BRANCH_NAME_CHARGE_HOLDER={record[0][21]}&BUILDING_NAME_CHARGE_HOLDER=&STREET_NAME_CHARGE_HOLDER=&VILLAGE_CHARGE_HOLDER=&TALUKA_CHARGE_HOLDER=&CITY_CHARGE_HOLDER=&DISTRICT_CHARGE_HOLDER={record[0][22]}&STATE_CHARGE_HOLDER={record[0][23] if int(record[0][23])>9 else F'0{record[0][23]}'}&PINCODE_CHARGE_HOLDER={record[0][24]}&TIPVALUE_CHARGE_HOLDER_BRANCH_TYPE=&TIPVALUE_CHARGE_HOLDER_BRANCH_CODE=&TIPVALUE_PINCODE_CHARGE_HOLDER=&TIPVALUE_STATE_CHARGE_HOLDER=&TIPVALUE_DISTRICT_CHARGE_HOLDER=&TIPVALUE_CITY_CHARGE_HOLDER=&TIPVALUE_TALUKA_CHARGE_HOLDER=&TIPVALUE_VILLAGE_CHARGE_HOLDER=&TIPVALUE_STREET_NAME_CHARGE_HOLDER=&TIPVALUE_BUILDING_NAME_CHARGE_HOLDER=&TIPVALUE_BRANCH_NAME_CHARGE_HOLDER=&FORCEFUL_INSERT_FLAG=N&ASSET_TYPE=01&SI_TYPE=2&OTHER_SI_TYPE=&ASSET_NATURE_OF_PROPERTY=0{str(record[0][1])[-1:]}&ASSET_SURVEY_NUMBER=&ASSET_PLOT_NUMBER={record[0][2]}&ASSET_CARPET_AREA={record[0][3]}&ASSET_CARPET_AREA_UNIT={unit}&ASSET_HOUSE_NUMBER={record[0][5]}&ASSET_FLOOR_NUMBER=&ASSET_BUILDING_NUMBER=&ASSET_BUILDING_NAME=&ASSET_BLOCK_NUMBER=&ASSET_STREET_NAME=&ASSET_LOCALITY={record[0][6]}&ASSET_STAGE_SECTOR_WARD=&ASSET_LANDMARK=&ASSET_LATITUDE_VAL=&ASSET_LONGITUDE_VAL=&ASSET_VILLAGE=&ASSET_TOWN=&ASSET_TALUKA=&ASSET_DISTRICT={record[0][7]}&ASSET_STATE={record[0][8] if int(record[0][8])>9 else F'0{record[0][8]}'}&ASSET_PINCODE={record[0][9]}&ASSET_NORTH_BOUNDARY=&ASSET_SOUTH_BOUNDARY=&ASSET_EAST_BOUNDARY=&ASSET_WEST_BOUNDARY=&TIPVALUE_ASSET_NATURE_OF_PROPERTY=&TIPVALUE_ASSET_SURVEY_NUMBER=&TIPVALUE_ASSET_PLOT_NUMBER=&TIPVALUE_ASSET_CARPET_AREA=&TIPVALUE_ASSET_CARPET_AREA_UNIT=&TIPVALUE_ASSET_HOUSE_NUMBER=&TIPVALUE_ASSET_FLOOR_NUMBER=&TIPVALUE_ASSET_BUILDING_NAME=&TIPVALUE_ASSET_BLOCK_NUMBER=&TIPVALUE_ASSET_STREET_NAME=&TIPVALUE_ASSET_LOCALITY=&TIPVALUE_ASSET_STAGE_SECTOR_WARD=&TIPVALUE_ASSET_LANDMARK=&TIPVALUE_ASSET_LATITUDE_VAL=&TIPVALUE_ASSET_LONGITUDE_VAL=&TIPVALUE_ASSET_VILLAGE=&TIPVALUE_ASSET_TOWN=&TIPVALUE_ASSET_TALUKA=&TIPVALUE_ASSET_DISTRICT=&TIPVALUE_ASSET_STATE=&TIPVALUE_ASSET_PINCODE=&TIPVALUE_ASSET_NORTH_BOUNDARY=&TIPVALUE_ASSET_SOUTH_BOUNDARY=&TIPVALUE_ASSET_WEST_BOUNDARY=&TIPVALUE_ASSET_EAST_BOUNDARY=&TIPVALUE_ASSET_BUILDING_NUMBER=&TIPVALUE_ASSET_TYPE=&SYSDATE={date_format(datetime.datetime.now())}&EVALUATED_PRICE_OF_ASSET={record[0][25]}&TIPVALUE_EVALUATED_PRICE_OF_ASSET=&CHARGE_CREATION_DATE={date_format(record[0][26])}&TIPVALUE_CHARGE_CREATION_DATE=&"


    st4 = ""
    for line in record[0:1]:
        dtype = "Sale+deed" if "sale" in str(line[27]).lower() else "Lease+deed" if "lease" in str(
            line[27]).lower() else "Award" if "award" in str(line[27]).lower() else "Others"
        dyo = line[28] if dtype == "Others" else ""

        st4 += F"DOCUMENT_TYPE={dtype}&DOCUMENT_TYPE_OTHERS={dyo}&DOCUMENT_NUMBER={line[29]}&DOCUMENT_DATE={date_format(line[30])}&DOCUMENT_SUB_REGISTRAR={line[31]}&DOCUMENT_TALUKA=&DOCUMENT_DISTRICT=&DOCUMENT_STATE={line[32] if int(line[32])>9 else F'0{line[32]}'}&DOCUMENT_PINCODE={line[33]}&TIPVALUE_DOCUMENT_DATE=&TIPVALUE_DOCUMENT_STATE=&TIPVALUE_DOCUMENT_DISTRICT=&TIPVALUE_DOCUMENT_TALUKA=&TIPVALUE_DOCUMENT_SUB_REGISTRAR=&TIPVALUE_DOCUMENT_NUMBER=&TIPVALUE_DOCUMENT_TYPE=&TIPVALUE_DOCUMENT_PINCODE=&"


    st5 = F"TOTAL_SECURED_AMOUNT={record[0][36]}&"
    st6 = ""
    for line in record[0:1]:
        nof = 'Demand+Loan' if 'demand loan' in str(line[34]).lower() else 'Term+Loan' if 'term loan' in str(
            line[34]).lower() else 'LAP' if 'lap' in str(line[34]).lower() else 'Cash+Credit' if 'cash credit' in str(
            line[34]).lower() else 'Overdraft' if 'overdraft' in str(line[34]).lower() else 'LC' if 'lc' in str(
            line[34]).lower() else 'BG' if 'bg' in str(line[34]).lower() else 'Bills' if 'bills' in str(
            line[34]).lower() else 'Derivatives' if 'derivatives' in str(
            line[34]).lower() else 'Export+Packing+Credit' if 'export packing credit' in str(
            line[34]).lower() else 'Foreign+Usance+Bills+Discounted' if 'foreign usance bills discounted' in str(
            line[34]).lower() else 'Foreign+Bills+Purchased' if 'foreign bills purchased' in str(line[34]).lower() else 'Post+Shipment+Credit+in+Foreign+Currency' if 'post shipment credit in foreign currency' in str(
            line[34]).lower() else 'ECB' if 'ecb' in str(line[34]).lower() else ""

        st6 += F"NATURE_OF_FACILITY={nof}&LOAN_ACCOUNT_NUMBER={line[35]}&SECURED_AMOUNT_ON_LOAN={line[36]}&LOAN_DATE={date_format(line[37])}&RATE_OF_INTEREST={line[38]}&REPAYMENT_PERIOD={line[39]}&EXTENT_AND_OPERATION=&OTHER_INFO=&TIPVALUE_NATURE_OF_FACILITY=&TIPVALUE_OTHER_INFO=&TIPVALUE_EXTENT_AND_OPERATION=&TIPVALUE_REPAYMENT_PERIOD=&TIPVALUE_RATE_OF_INTEREST=&TIPVALUE_LOAN_ACCOUNT_NUMBER=&TIPVALUE_SECURED_AMOUNT_ON_LOAN=&TIPVALUE_LOAN_DATE=&"


    st7 = F"NO_OF_BORROWERS={len(record)}&NO_OF_TPMS=0&NO_OF_DOCUMENTS=1&NO_OF_LOANS=1"


    final_str = st1+st2+st3+st4+st5+st6+st7
    final = urllib.parse.quote_plus(final_str,safe="&=+")

    return final


def start_process(record, driver, maker_id, password):
    if len(record) < 1:
        return "Fail", "ZERO RECORD"

    # driver = webdriver.Chrome("/usr/bin/chromedriver")
    driver.maximize_window()
    # driver = webdriver.Chrome(os.path.join(crm_path,"chromedriver.exe"))
    driver.get("https://www.cersai.org.in/CERSAI/")

    wait = ui.WebDriverWait(driver, 10)
    wait.until(page_is_loaded)

    driver.find_element_by_name('USERNAME').send_keys(maker_id)
    driver.find_element_by_name('PASSWORD').send_keys(password)

    captcha = eval(driver.find_element_by_id("CAPTCHA_QUESTION").get_attribute("value"))

    driver.find_element_by_name('CAPTCHA_ANSWER').send_keys(captcha)

    driver.implicitly_wait(3)
    driver.find_element_by_name('SUBMITLOGIN').click()

    wait = ui.WebDriverWait(driver, 10)
    wait.until(page_is_loaded)
    driver.find_element_by_xpath('//*[@id="sidebar"]/form/ul/li[3]/ul/li[1]/a')
    # driver.find_element_by_xpath('//*[@id="sidebar"]/form/ul/li[3]/ul/li[1]/a').click()
    #
    # wait = ui.WebDriverWait(driver, 10)
    # wait.until(page_is_loaded)
    #
    # select = Select(driver.find_element_by_id('ASSET_TYPE'))
    # select.select_by_visible_text('Immovables')
    #
    # select = Select(driver.find_element_by_id('SI_TYPE'))
    # select.select_by_visible_text('Registered Mortgage')
    #
    # driver.find_element_by_id('JOINT_N').click()

    # select = Select(driver.find_element_by_id('NATURE_OF_PROPERTY'))
    # select.select_by_visible_text('Residential Plot')
    # select.select_by_value("0" + str(int(record[0][1])))
    try:
        cc = driver.get_cookies()
        sid = cc[0].get("value")
    except:
        return "Fail", "No Cookies Found", 111
    final_data = get_data(record, sid, maker_id)
    print(final_data)
    token_data,statuscode = submit_req(final_data, sid)
    print(statuscode)

    if "Request Accepted Successfully" in token_data:
        tmp = re.findall("\d{6}\d+", str(token_data))
        rid = F"#{tmp[0]}"
        print("Successfull")
        time.sleep(5)
        return "Success", rid, statuscode
    else:
        print("unsuccessfull")
        time.sleep(5)
        return "Fail", token_data, 111


def records_left(f_path):
    fname = os.path.join(f_path, "main_file.obj")
    with open(fname, "rb") as queue_save_file:
        try:
            data = pickle.load(queue_save_file)
        except:
            pass
    totalele = []
    for c in data:
        totalele.append(c[0])
    return len(set(totalele)) + 1


def get_record(f_path):
    files = sorted(os.listdir(f_path))
    for i, f in enumerate(files):
        if not ".obj" in str(f):  # Removing non obj files
            files.pop(i)

    # if len(files)<1:
    #     al = pyautogui.alert("Folder name 'Queue_to_be_processed' has no files with '.obj' extention. Please move files into the folder to process.\n\n\nPress OK to Abort. ")
    #     if al == "OK":
    #         sys.exit()

    m_file = ""
    for i, x in enumerate(files):
        if "main_file" in str(x):
            m_file = x
            files.pop(i)
            break
    if "main_file" in str(m_file):
        try:
            fname = os.path.join(f_path, m_file)
        except:
            al = pyautogui.alert(
                "Folder name 'Queue_to_be_processed' does not contain any file named as main_file.\n\n\nPress OK to Abort. ")
            if al == "OK":
                sys.exit()
    else:
        copyfile(os.path.join(os.getcwd(), "Files", "util", "main_file.obj"), os.path.join(f_path, "main_file.obj"))
        fname = os.path.join(f_path, "main_file.obj")

    record = []
    with open(fname, "rb") as queue_save_file:
        try:
            data = pickle.load(queue_save_file)
        except EOFError:
            if len(files) >= 1:
                append_obj_files(fname, os.path.join(f_path, str(files[0])))

                with open(fname, "rb") as queue_save_file:
                    data = pickle.load(queue_save_file)

    if len(data) < 1 and len(files) >= 1:
        append_obj_files(fname, os.path.join(f_path, str(files[0])))

        with open(fname, "rb") as queue_save_file:
            data = pickle.load(queue_save_file)

    if len(data) >= 1:
        temp = data.popleft()
        record.append(temp)
        try:
            while temp[0] == data[0][0] or numpy.isnan(data[0][0]):
                record.append(data.popleft())
        except:
            pass
        with open(fname, "wb+") as queue_save_file:
            pickle.dump(data, queue_save_file)
    if len(record) > 0:
        return record
    else:
        pyautogui.alert("No more records left. Please add more files", "Alert")
        return []


def excel_to_queue():
    filename = "/home/samit/Downloads/Cersai Data-Jan2019-robot-240 cases.xlsx"
    # filename = "Cersai Data_Jan2019_RPA_Swati.xlsx"
    df = pandas.read_excel(filename)
    Q = collections.deque()
    np_df = df.values
    i = 0
    for x in np_df:
        Q.append(x)
        i += 1
    try:
        while numpy.isnan(Q[len(Q) - 1][10]):
            Q.pop()
            i -= 1
    except:
        pass

    with open("xls_to_queue.obj", "wb+") as queue_save_file:
        pickle.dump(Q, queue_save_file)


def append_obj(filename, data):
    if os.path.isfile(filename):
        with open(filename, "r+b") as load_file:
            loaded_data = pickle.load(load_file)
        for itm in data:
            loaded_data.append(itm)
        with open(filename, "wb+") as queue_save_file:
            pickle.dump(loaded_data, queue_save_file)
    else:
        with open(filename, "wb+") as queue_save_file:
            pickle.dump(data, queue_save_file)


def append_obj_files(file1, file2):
    print("appending file {} and {}".format(file1, file2))
    with open(file1, "r+b") as load_file:
        try:
            loaded_data1 = pickle.load(load_file)
        except EOFError:
            loaded_data1 = ""

    with open(file2, "r+b") as load_file:
        try:
            loaded_data2 = pickle.load(load_file)
        except EOFError:
            loaded_data2 = ""

    if len(loaded_data1) < len(loaded_data2):
        for itm in loaded_data1:
            loaded_data2.append(itm)

        with open(file1, "wb+") as queue_save_file:
            pickle.dump(loaded_data2, queue_save_file)
    else:
        for itm in loaded_data2:
            loaded_data1.append(itm)

        with open(file1, "wb+") as queue_save_file:
            pickle.dump(loaded_data1, queue_save_file)

    remove_files(file2)


def remove_files(file_path):
    os.remove(file_path)


def obj_to_excel(fpath, exlpath):
    files = os.listdir(fpath)
    for f in files:
        with open(f, "r+b") as load_file:
            loaded_data1 = pickle.load(load_file)

        wb = Workbook()
        ws = wb.active

        ws.append(["sr. no.", "Nature of Property", "Plot No / Survey No", " CARPET/BUILD UP AREA           ",
                   "Square Feet/Square Meter", "House No/Flat No/Bunglow No", "Locality", "District",
                   "State / Union Territory", "PIN Code ", "BORROWER Type (INDIVIDUAL)", "ASSET OWNER (YES)/(NO)",
                   "Title ",
                   "Name", "Date of Birth", "Individual Borrower PAN", "Plot No/Survery No/House-Flat No", "State ",
                   "District", "PIN Code ", "Branch Code", "Branch Name", "State", "District", "PIN Code ",
                   "Evaluated Price of Asset*", "Security Interest creation date within bank", "Document Type",
                   "SPECIFY DOCUMENT TYPE", "Title Document No", "Document date", "Sub Registrar", "State", "PIN Code ",
                   "Nature of facility (TERM LOAN)", "Loan Account Number", "Loan Amount", "Date of Disbursement",
                   "Rate of Interest", "Repayment Period (IN MONTH)"])

        for i in loaded_data1:
            ws.append(list(i))
        xlname = "{}.xlsx".format(f.split(".")[0])
        wb.save(os.path.join(exlpath, xlname))


#
#
# def to_csv(filepath, data, status, mispath=None, rid=None):
#
#     csvHeader1 = ["Sr. No.", "Nature of Property", "Plot No", "Built-up Area", "Square Feet/Square Meter", "House/Flat/Shop No.",
#      "Locality", "District", "State / UT", "PIN Code ", "BORROWER Type", "ASSET OWNER (YES)/(NO)", "Title ", "Name",
#      "Date of Birth", "Plot No/Survery No/Flat/House No.", "District", "State ", "PIN Code ", "Individual Borrower PAN",
#      "Branch Code", "Branch Name", "District", "State", "PIN Code ", "Evaluated Price of Asset*",
#      "Security Interest creation date within bank", "Document Type", "SPECIFY DOCUMENT TYPE", "Title Document No",
#      "Document date", "Sub Registrar", "State", "PIN Code ", "Nature of facility", "Loan Account Number", "Loan Amount",
#      "Date of Disbursement", "Rate of Interest", "Repayment Period (in months)", "Token", "Status", "Date"]
#     csvHeader2 =["Sr. No.", "Nature of Property", "Plot No", "Built-up Area", "Square Feet/Square Meter", "House/Flat/Shop No.",
#      "Locality", "District", "State / UT", "PIN Code ", "BORROWER Type", "ASSET OWNER (YES)/(NO)", "Title ", "Name",
#      "Date of Birth", "Plot No/Survery No/Flat/House No.", "District", "State ", "PIN Code ", "Individual Borrower PAN",
#      "Branch Code", "Branch Name", "District", "State", "PIN Code ", "Evaluated Price of Asset*",
#      "Security Interest creation date within bank", "Document Type", "SPECIFY DOCUMENT TYPE", "Title Document No",
#      "Document date", "Sub Registrar", "State", "PIN Code ", "Nature of facility", "Loan Account Number", "Loan Amount",
#      "Date of Disbursement", "Rate of Interest", "Repayment Period (in months)", "Status","Error Message", "Date"]
#     if os.path.isfile(filepath):
#         mode = "a"
#     else:
#         mode = "a+"
#
#     if rid and status == "Success":
#         if mispath:
#             if os.path.isfile(mispath):
#                 mm = "a"
#             else:
#                 mm = "a+"
#             with open(mispath, mm, newline='') as misfile:
#                 csvWriter = csv.writer(misfile, delimiter=',', quotechar="\"", quoting=csv.QUOTE_MINIMAL)
#                 if mode == "a+":
#                     csvWriter.writerow(["FinnOne No.", "Customer Name", "State", "Branch", "Disbursement Date", "Maker ID", "Maker Date","Token"])
#                 csvWriter.writerow([data[0][35], data[0][13], data[0][17], data[0][21], data[0][37], maker_id, get_date(), rid])
#
#         with open(filepath, mode, newline='') as csvFile:
#             csvWriter = csv.writer(csvFile, delimiter=',', quotechar="\"", quoting=csv.QUOTE_MINIMAL)
#             if mode == "a+":
#                 csvWriter.writerow(csvHeader1)
#
#             for i, x in enumerate(data):
#                 print(x)
#                 x = list(x)
#                 if i == 0:
#                     x.extend([rid, "Success", get_date()])
#                     csvWriter.writerow(x)
#                 else:
#                     csvWriter.writerow(x)
#     else:
#         with open(filepath, mode, newline='') as csvFile:
#             csvWriter = csv.writer(csvFile, delimiter=',', quotechar="\"", quoting=csv.QUOTE_MINIMAL)
#             if mode == "a+":
#                 csvWriter.writerow(csvHeader2)
#
#             for i, x in enumerate(data):
#                 print(x)
#                 x = list(x)
#                 if i == 0:
#                     x.extend(["Fail",status,get_date()])
#                     csvWriter.writerow(x)
#                 else:
#                     csvWriter.writerow(x)
#

def to_excel(filepath, data, status, mispath=None, rid=None, maker_id=None):
    Header1 = ["Sr. No.", "Nature of Property", "Plot No", "Built-up Area", "Square Feet/Square Meter",
               "House/Flat/Shop No.",
               "Locality", "District", "State / UT", "PIN Code ", "BORROWER Type", "ASSET OWNER (YES)/(NO)", "Title ",
               "Name",
               "Date of Birth", "Plot No/Survery No/Flat/House No.", "District", "State ", "PIN Code ",
               "Individual Borrower PAN",
               "Branch Code", "Branch Name", "District", "State", "PIN Code ", "Evaluated Price of Asset*",
               "Security Interest creation date within bank", "Document Type", "SPECIFY DOCUMENT TYPE",
               "Title Document No",
               "Document date", "Sub Registrar", "State", "PIN Code ", "Nature of facility", "Loan Account Number",
               "Loan Amount",
               "Date of Disbursement", "Rate of Interest", "Repayment Period (in months)", "Queue Reference Number",
               "Security Interest ID", "Asset ID", "Token", "Status", "Date"]
    Header2 = ["Sr. No.", "Nature of Property", "Plot No", "Built-up Area", "Square Feet/Square Meter",
               "House/Flat/Shop No.",
               "Locality", "District", "State / UT", "PIN Code ", "BORROWER Type", "ASSET OWNER (YES)/(NO)", "Title ",
               "Name",
               "Date of Birth", "Plot No/Survery No/Flat/House No.", "District", "State ", "PIN Code ",
               "Individual Borrower PAN",
               "Branch Code", "Branch Name", "District", "State", "PIN Code ", "Evaluated Price of Asset*",
               "Security Interest creation date within bank", "Document Type", "SPECIFY DOCUMENT TYPE",
               "Title Document No",
               "Document date", "Sub Registrar", "State", "PIN Code ", "Nature of facility", "Loan Account Number",
               "Loan Amount",
               "Date of Disbursement", "Rate of Interest", "Repayment Period (in months)", "Queue Reference Number",
               "Security Interest ID", "Asset ID", "Status", "Error Message", "Date"]

    if rid and status == "Success":
        if mispath:
            if not os.path.isfile(mispath):
                misbook = Workbook()
                missheet = misbook.active
                missheet.append(
                    ["FinnOne No.", "Customer Name", "State", "Branch", "Disbursement Date", "Maker ID", "Maker Date",
                     "Token"])
            else:
                misbook = openpyxl.load_workbook(mispath)
                missheet = misbook.active

            missheet.append(
                [data[0][35], data[0][13], data[0][17], data[0][21], data[0][37], maker_id, get_date(), rid])
            misbook.save(mispath)

        if not os.path.isfile(filepath):
            book = Workbook()
            sheet = book.active
            sheet.append(Header1)
        else:
            book = openpyxl.load_workbook(filepath)
            sheet = book.active

        for i, x in enumerate(data):
            x = list(x)
            if i == 0:
                x.extend([rid, "Success", get_date()])
            sheet.append(x)
        book.save(filepath)
    else:
        if not os.path.isfile(filepath):
            book = Workbook()
            sheet = book.active
            sheet.append(Header2)
        else:
            book = openpyxl.load_workbook(filepath)
            sheet = book.active

        for i, x in enumerate(data):
            x = list(x)
            if i == 0:
                x.extend(["Fail", str(status), get_date()])
            sheet.append(x)
        book.save(filepath)


def make_dir(*paths):
    # Creates all required directories if not present, mentioned in the path.
    for pt in paths:
        if not (os.path.isdir(pt)):
            try:
                os.makedirs(pt, mode=0o777, exist_ok=True)
            except OSError as exception:
                if exception.errno != errno.EEXIST:
                    raise


def get_date():
    dt = datetime.datetime.now()
    date = dt.strftime("%d%m%Y")
    return str(date)


def mailer():
    print("sending mail")
    fromaddr = "samit.pawar94@gmail.com "
    toaddr = ["sameer.kulkarni@sequelstring.com"]
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = ", ".join(toaddr)
    msg['Subject'] = "Issue Alert"
    body = """Hello there,
    this  is  an issue in.
    Please Check."""
    msg.attach(MIMEText(body, 'plain'))
    s = smtplib.SMTP('smtp.gmail.com', 587)
    s.starttls()
    s.login(fromaddr, "spsingh94")
    text = msg.as_string()
    s.sendmail(fromaddr, toaddr, text)
    s.quit()
    print("mail sent")


def create_rotating_logger(path):
    """
    Creates a rotating log
    """
    logger = logging.getLogger("Maker Logs")
    # Assigning Level
    logger.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

    # add a rotating handler
    handler = RotatingFileHandler(path, mode='a+', maxBytes=2*1048576, backupCount=20)
    handler.setFormatter(formatter)
    logger.addHandler(handler)

    return logger


#
#
#
# QUEUE_TBP_PATH = os.path.join(os.getcwd(), "Files", "Queue_to_be_processed")
# data = get_record(QUEUE_TBP_PATH)
#
# get_data(data,"SESSION ID","F010050159")
#
#
# mailer()
#
#
#
# tt = pyautogui.locateOnScreen(img,)
# i = 1
# while not tt and i<15:
#     time.sleep(1)
#     tt = pyautogui.locateOnScreen(img,)
#     i += 1
#
