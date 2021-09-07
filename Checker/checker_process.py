import datetime
import subprocess
import sys
import threading
import time
import pandas
import os
import re
import openpyxl
import pyautogui
import autoit
from pubsub import pub
import wx
import errno
import base64
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import ui
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException


class LoginDialog(wx.Dialog):
    """
    Class to define login dialog
    """

    # ----------------------------------------------------------------------
    def __init__(self):
        font = wx.SystemSettings.GetFont(wx.SYS_SYSTEM_FONT)
        font.SetPointSize(20)
        self.flag = 0
        """Constructor"""
        wx.Dialog.__init__(self, None, title="CHECKER LOGIN")
        # user info
        user_sizer = wx.BoxSizer(wx.HORIZONTAL)
        user_lbl = wx.StaticText(self, label="Checker ID:")
        user_lbl.SetFont(font)
        user_sizer.Add(user_lbl, 0, wx.ALL | wx.CENTER, 5)
        self.user = wx.TextCtrl(self)
        user_sizer.Add(self.user, 0, wx.ALL, 5)

        # pass info
        p_sizer = wx.BoxSizer(wx.HORIZONTAL)
        p_lbl = wx.StaticText(self, label="Password:  ")
        p_lbl.SetFont(font)
        p_sizer.Add(p_lbl, 0, wx.ALL | wx.CENTER, 5)
        self.password = wx.TextCtrl(self, style=wx.TE_PASSWORD | wx.TE_PROCESS_ENTER)
        p_sizer.Add(self.password, 0, wx.ALL, 5)

        main_sizer = wx.BoxSizer(wx.VERTICAL)
        self.ulabel = wx.StaticText(self, -1, "Please Enter Checker ID and Password to Continue")
        main_sizer.Add(user_sizer, 0, wx.ALL, 10)
        main_sizer.Add(p_sizer, 0, wx.ALL, 10)
        main_sizer.Add(self.ulabel, 0, wx.ALL, 10)

        # btn = wx.Button(self, label="Login")
        btn = wx.Button(self, 1, label="LOGIN", pos=(140, 150), size=(120, 40))
        btn.SetBackgroundColour("#00A6FF")
        btn.SetFont(font)
        btn.SetForegroundColour("#FAFAFA")

        btn.Bind(wx.EVT_BUTTON, self.onLogin)

        # main_sizer.Add(btn, 0, wx.ALL | wx.CENTER, 5)
        self.SetSizer(main_sizer)

    # ----------------------------------------------------------------------
    def onLogin(self, event):

        """
        Check credentials and login  200105271134
        """
        creds = {"F010200062": "mrhfl@1", "F010200061": "Kumar@123", "F010200002": "chalke@2", "F010200078": "mrhfl@1",
                 "F010200079": "mrhfl@1"}
        self.user_name = self.user.GetValue()
        self.user_password = self.password.GetValue()
        try:
            getuser = self.user_name
        except KeyError:
            getuser = None
            self.ulabel.SetLabelText("Invalid Checker ID")

        if getuser and creds.get(getuser):
            if not self.user_password or creds[getuser] != self.user_password:
                self.ulabel.SetLabelText("Invalid Password")
            else:
                self.flag = 1
                self.ulabel.SetLabelText("You are now logged in!")
                pub.sendMessage("frameListener", message="show")
                self.Destroy()
        else:
            self.ulabel.SetLabelText("Invalid Checker ID")


class Checker_Process(wx.Frame):
    def __init__(self, parent, id):
        wx.Frame.__init__(self, parent, id, "CERSAI", size=(420, 350))
        self.panel = wx.Panel(self)

        pub.subscribe(self.myListener, "frameListener")
        # Ask user to login
        self.dlg = LoginDialog()

        self.dlg.ShowModal()

        if self.dlg.flag != 1:
            sys.exit()
        try:
            image_file = os.path.join(util_path, 'sslogo2.jpeg')
            bmp1 = wx.Image(image_file, wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            self.bitmap1 = wx.StaticBitmap(self.panel, -1, bmp1, (0, 0))
        except:
            pass

        self.label = wx.StaticText(self.panel, -1, "Please Enter Maker ID to Start", (10, 290))
        self.start_button = wx.Button(self.panel, 1, label="START", pos=(60, 220), size=(120, 60))
        self.stop_button = wx.Button(self.panel, label="STOP", pos=(220, 220), size=(120, 60))
        self.Bind(wx.EVT_BUTTON, self.start_thread, self.start_button)
        self.idtext = wx.TextCtrl(self.panel, pos=(140, 170), size=(200, 20))
        self.idlabel = wx.StaticText(self.panel, -1, "Maker ID:", (60, 172))

        self.Bind(wx.EVT_BUTTON, self.stop_f, self.stop_button)
        font = wx.SystemSettings.GetFont(wx.SYS_SYSTEM_FONT)

        font.SetPointSize(20)
        self.start_button.SetBackgroundColour("#0B0B3B")
        self.start_button.SetFont(font)
        self.start_button.SetForegroundColour("#FAFAFA")
        self.stop_button.SetBackgroundColour("#0B0B3B")
        self.stop_button.SetFont(font)
        self.stop_button.SetForegroundColour("#FAFAFA")

        # self.label.SetForegroundColour((255, 0, 0))  # set text color

    def myListener(self, message, arg2=None):
        """
        Show the frame
        """
        self.Show()

    def page_is_loaded(self, driver):
        return driver.find_element_by_tag_name("body") != None

    def stop_f(self, event):
        try:
            self.driver.quit()
        except:
            pass
        sys.exit()

    def start_thread(self, event):
        ids = ["F010200015", "F010200017", "F010200050", "F010200057", "F010200058", "F010200059", "F010200070",
               "F010200071", "F010200072", "F010200073", "F010200080", "F010200081"]
        self.makerid = self.idtext.GetValue()
        if self.idtext.GetValue():
            if str(self.idtext.GetValue()) not in ids:
                self.label.SetForegroundColour((255, 0, 0))
                self.label.SetLabelText("Please Enter a Valid Maker ID")
            else:
                self.t1 = threading.Thread(target=self.start_process)
                self.t1.setDaemon(True)
                self.t1.start()
                self.label.SetForegroundColour((0, 0, 0))
                self.label.SetLabelText("Started")
        else:
            self.label.SetForegroundColour((255, 0, 0))  # set text color
            self.label.SetLabelText("Enter the Maker id First")
            # self.label.SetBackgroundColour((0, 0, 255))  # set text back color

    def start_process(self):
        # self.driver = webdriver.Firefox(executable_path=cpath)
        # self.driver = webdriver.Ie(cpath)
        self.driver = webdriver.Chrome(cpath)
        # self.driver = webdriver.Chrome("/usr/bin/chromedriver")
        try:
            self.driver.maximize_window()
        except:
            pass
        self.driver.get("https://www.cersai.org.in/CERSAI/")

        wait = ui.WebDriverWait(self.driver, 10)
        wait.until(self.page_is_loaded)
        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.NAME,"USERNAME")))

        self.driver.find_element_by_name('USERNAME').send_keys(self.dlg.user_name)
        self.driver.implicitly_wait(2)
        self.driver.find_element_by_name('PASSWORD').send_keys(self.dlg.user_password)
        self.driver.implicitly_wait(2)
        captcha = eval(self.driver.find_element_by_id("CAPTCHA_QUESTION").get_attribute("value"))

        self.driver.find_element_by_name('CAPTCHA_ANSWER').send_keys(captcha)

        self.driver.implicitly_wait(3)

        self.driver.find_element_by_name('SUBMITLOGIN').click()

        wait = ui.WebDriverWait(self.driver, 10)
        wait.until(self.page_is_loaded)
        self.driver.find_element_by_id("CB").click()
        time.sleep(2)
        self.driver.implicitly_wait(3)
        self.driver.find_element_by_id("SUBMIT").click()
        time.sleep(2)
        autoit.control_focus("Signing Wizard", "SysHeader321")
        autoit.control_click("Signing Wizard", "SysHeader321")
        pyautogui.hotkey("down")
        autoit.control_focus("Signing Wizard", "Button1")
        autoit.control_click("Signing Wizard", "Button1")
        wait = ui.WebDriverWait(self.driver, 10)
        wait.until(self.page_is_loaded)
        WebDriverWait(self.driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="sidebar"]/form/ul/li[1]/ul/li[2]/a'))).click()

        wait = ui.WebDriverWait(self.driver, 10)
        wait.until(self.page_is_loaded)
        time.sleep(2)
        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.ID,"MAKER_USER_ID")))
        self.driver.find_element_by_id("MAKER_USER_ID").send_keys(self.makerid)
        sl = Select(self.driver.find_element_by_id("PROCESS_ID"))
        sl = Select(self.driver.find_element_by_id("PROCESS_ID"))
        sl.select_by_visible_text("Add Security Interest")
        self.driver.find_element_by_id("SUBMIT").click()

        wait = ui.WebDriverWait(self.driver, 15)
        wait.until(self.page_is_loaded)
        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="content"]/div[1]/div/form/center[2]/div/table/tbody/tr[1]/th[1]')))
        WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="content"]/div[1]/div/form/center[2]/div/table/tbody/tr[1]/th[1]'))).click()
        time.sleep(3)
        fname = os.path.join(TOKEN_PATH, "MIS_FILE.xlsx")
        df = pandas.read_excel(fname)
        allvals = [int(x.replace("#", "")) for x in df['Token'].tolist()]
        df = []
        WebDriverWait(self.driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="content"]/div[1]/div/form/center[2]/div/table/tbody/tr')))
        ActionChains(self.driver).key_down(Keys.END).key_down(Keys.END).key_up(Keys.END).perform()
        time.sleep(2)
        all_ele = self.driver.find_elements_by_xpath('//*[@id="content"]/div[1]/div/form/center[2]/div/table/tbody/tr')
        print("all element length = ", len(all_ele))

        finaldt = []
        i = 1
        while len(all_ele) > 1 and i <= len(all_ele):
            # for i in range(len(all_ele)):
            tm1 = int(time.time())
            rec = []
            try:
                x = all_ele[i]
            except:
                break
            i += 1
            dd = (x.text).split("\n")
            tkn = str(dd[0]).split()[0]
            print("checking token:", tkn)
            print("i", i)
            try:
                igdf = pandas.read_excel(IGNORE_FILE)
                ignorevals = [int(x.replace("#", "")) for x in igdf['Skipped QRN'].tolist()]
            except:
                ignorevals = []
            if int(tkn) in allvals and int(tkn) not in ignorevals:
                print("Processing token:", tkn)
                bname = dd[0].split(" ", 1)[1].split("Immovable")[0]
                mdate = dd[-1].split()[0]
                if len(tkn) > 10:
                    time.sleep(2)
                    self.driver.find_element_by_id(tkn).click()
                    time.sleep(2)
                    self.driver.find_element_by_id("CHECK_RECORD").click()
                    try:
                        time.sleep(2)
                        errormsg = self.driver.find_element_by_id("MESSAGE_ERROR").text
                        if "Please select a record" in errormsg:
                            time.sleep(2)
                            self.driver.find_element_by_id(tkn).click()
                            time.sleep(1)
                            self.driver.find_element_by_id("CHECK_RECORD").click()
                            # errormsg = self.driver.find_element_by_id("MESSAGE_ERROR").text
                    except Exception as e:
                        print("Error:",e)
                        pass

                    time.sleep(3)
                    try:
                        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.ID,"btnSubmit")))
                        self.driver.find_element_by_id('btnSubmit').click()
                    except Exception as e:
                        print(e)
                        pass

                    time.sleep(4)
                    for x in range(1, 8):
                        if x != 5:
                            WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.ID, f"TAB_BUTTON_{x}")))
                            self.driver.find_element_by_id("TAB_BUTTON_{}".format(x)).click()
                            time.sleep(1)
                            if x == 7:
                                time.sleep(1)
                                WebDriverWait(self.driver, 20).until(
                                    EC.element_to_be_clickable((By.ID, "LOAN_ACCOUNT_NUMBER_1")))
                                lacn = self.driver.find_element_by_id("LOAN_ACCOUNT_NUMBER_1").get_attribute('value')
                                dod = self.driver.find_element_by_id("LOAN_DATE_1").get_attribute('value')

                            if x == 4:
                                time.sleep(2)
                                WebDriverWait(self.driver, 20).until(
                                    EC.element_to_be_clickable((By.ID, f"BRANCH_NAME_CHARGE_HOLDER")))

                                branch = self.driver.find_element_by_id("BRANCH_NAME_CHARGE_HOLDER").get_attribute(
                                    'value')
                                state = self.driver.find_element_by_id("STATE_CHARGE_HOLDER").get_attribute('value')
                            self.driver.implicitly_wait(3)
                    try:
                        time.sleep(5)
                        WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.ID, f"CONFIRM")))
                        self.driver.find_element_by_id("CONFIRM").click()
                    except Exception as e:
                        err = "Unable to click on CONFIRM"
                        print("Error:",e)
                        print(err)
                    time.sleep(10)
                    wait.until(self.page_is_loaded)
                    try:
                        err = self.driver.find_element_by_id("MESSAGE_ERROR").text
                    except:
                        err = None
                    try:
                        success = self.driver.find_element_by_id("MESSAGE_SUCCESS").text
                    except:
                        success = None

                    try:
                        wrn = self.driver.find_element_by_id("MESSAGE_WARNING").text
                    except:
                        wrn = None

                    try:
                        inf = self.driver.find_element_by_id("MESSAGE_INFO").text
                    except:
                        inf = None
                    if not success and not err and not wrn and not inf:
                        try:
                            time.sleep(5)
                            WebDriverWait(self.driver, 10).until(
                                EC.element_to_be_clickable((By.ID, f"CONFIRM")))
                            self.driver.find_element_by_id("CONFIRM").click()
                        except Exception as e:
                            err = "Unable to click on CONFIRM"
                            print("Error:",e)
                            print(err)
                        time.sleep(10)
                        wait.until(self.page_is_loaded)
                        try:
                            err = self.driver.find_element_by_id("MESSAGE_ERROR").text
                        except:
                            err = None
                        try:
                            success = self.driver.find_element_by_id("MESSAGE_SUCCESS").text
                        except:
                            success = None

                        try:
                            wrn = self.driver.find_element_by_id("MESSAGE_WARNING").text
                        except:
                            wrn = None

                        try:
                            inf = self.driver.find_element_by_id("MESSAGE_INFO").text
                        except:
                            inf = None

                    c = re.compile(r'\d\d\d\d\d\d\d\d\d\d\d\d')
                    nums = c.findall(success)
                    if len(nums) > 2:
                        rec.extend([lacn, bname, state, branch, dod, self.idtext.GetValue(), mdate, self.dlg.user_name,
                                    self.get_date(), "", "First Checker Done", "", tkn, nums[1], nums[2], "", ""])
                    else:
                        rec.extend([lacn, bname, state, branch, dod, self.idtext.GetValue(), mdate, self.dlg.user_name,
                                    self.get_date(), "", "First Checker Done", "", tkn, "", "", "", ""])

                    if success:
                        rec.append(success)
                        self.create_excel(rec, os.path.join(SUCCESS_PATH,
                                                            F"Success_{self.get_date()}({self.dlg.user_name}).xlsx"))

                    elif err:

                        if "Error in writing to database." in str(err) or "No Transaction or Account Details returned." in str(err):
                            self.driver.find_element_by_id("REJECT").click()
                            print("rejected")
                            self.driver.find_element_by_id('REJECTION_COMMENTS').send_keys("Inappropriate data")
                            time.sleep(1)
                            self.driver.find_element_by_id("REJECT").click()
                            self.reject_excel(
                                [lacn, bname, dod, self.idtext.GetValue(), self.dlg.user_name, "REJECTED", f"#{int(tkn)}", self.get_date(), str(err)], os.path.join(REJECT_PATH, f"rejected_{self.get_date()}.xlsx"))
                        else:
                            rec.append("ERROR:" + str(err))
                            self.ignore_vals_excel([f"#{int(tkn)}"],IGNORE_FILE)
                            self.create_excel(rec, os.path.join(FAIL_PATH,
                                                            F"Fail_{self.get_date()}({self.dlg.user_name}).xlsx"))
                    elif wrn:
                        # ignorevals.append(int(tkn))
                        self.ignore_vals_excel([f"#{int(tkn)}"], IGNORE_FILE)
                        rec.append("WARNING" + str(wrn))
                        self.create_excel(rec, os.path.join(FAIL_PATH,
                                                            F"Fail_{self.get_date()}({self.dlg.user_name}).xlsx"))
                    elif inf:
                        # ignorevals.append(int(tkn))
                        self.ignore_vals_excel([f"#{int(tkn)}"],IGNORE_FILE)
                        rec.append("INFO" + str(inf))
                        self.create_excel(rec, os.path.join(FAIL_PATH,
                                                            F"Fail_{self.get_date()}({self.dlg.user_name}).xlsx"))
                    else:
                        # ignorevals.append(int(tkn))

                        self.ignore_vals_excel([f"#{int(tkn)}"], IGNORE_FILE)
                        rec.append("NO ELEMENT FOUND")
                        self.create_excel(rec, os.path.join(FAIL_PATH,
                                                            F"Fail_{self.get_date()}({self.dlg.user_name}).xlsx"))
                    
                    ActionChains(self.driver).key_down(Keys.END).key_down(Keys.END).key_up(Keys.END).perform()
                    WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable((By.ID, 'FYA_REDIRECTOR')))
                    time.sleep(3)

                    fyael = self.driver.find_element_by_id("FYA_REDIRECTOR")
                    fyael.send_keys("\n")

                        # try:
                        #     self.driver.find_element_by_id("FYA_REDIRECTOR").click()
                        # except Exception as e:
                        #     ActionChains(self.driver).key_down(Keys.END).key_down(Keys.END).key_up(Keys.END).perform()
                        #     self.driver.find_element_by_id("FYA_REDIRECTOR").click()
                        #     print(e)


                    print("Processed", tkn)
                    wait.until(self.page_is_loaded)
                    print("Please wait going back...")
                    try:
                        WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable(
                            (By.XPATH, '//*[@id="content"]/div[1]/div/form/center[2]/div/table/tbody/tr')))
                        ActionChains(self.driver).key_down(Keys.END).key_down(Keys.END).key_up(Keys.END).perform()
                        time.sleep(2)
                        all_ele = self.driver.find_elements_by_xpath(
                            '//*[@id="content"]/div[1]/div/form/center[2]/div/table/tbody/tr')
                        i = 1
                        print("all element length = ", len(all_ele))
                    except TimeoutException as e:
                        ActionChains(self.driver).key_down(Keys.END).key_down(Keys.END).key_up(Keys.END).perform()
                        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, 'FYA_REDIRECTOR')))
                        time.sleep(1)
                        fyael = self.driver.find_element_by_id("FYA_REDIRECTOR")
                        fyael.send_keys("\n")
                        WebDriverWait(self.driver, 15).until(EC.element_to_be_clickable(
                            (By.XPATH, '//*[@id="content"]/div[1]/div/form/center[2]/div/table/tbody/tr')))
                        ActionChains(self.driver).key_down(Keys.END).key_down(Keys.END).key_up(Keys.END).perform()
                        time.sleep(2)
                        all_ele = self.driver.find_elements_by_xpath(
                            '//*[@id="content"]/div[1]/div/form/center[2]/div/table/tbody/tr')
                        i = 1
                        print("all element length = ", len(all_ele))
                    except Exception:
                        time.sleep(7)
                        ActionChains(self.driver).key_down(Keys.END).key_down(Keys.END).key_up(Keys.END).perform()
                        time.sleep(2)
                        all_ele = self.driver.find_elements_by_xpath(
                            '//*[@id="content"]/div[1]/div/form/center[2]/div/table/tbody/tr')
                        print("all element length = ", len(all_ele))
                        i = 1
                print("time taken: ", int(time.time()) - tm1)
            else:
                print("Not Found")

        print("Completed")

    def reject_excel(self, record, fpath):
        if not os.path.isfile(fpath):
            book = openpyxl.Workbook()
            sheet = book.active
            sheet.append(["FinnOne No", "Customer Name", "Date of Disbursement","Maker ID", "Checker ID", "STATUS (Successful/Rejected)", "QRN",
                          "Date", "Error Message"])
        else:
            book = openpyxl.load_workbook(fpath)
            sheet = book.active

        sheet.append(record)
        book.save(fpath)

    def ignore_vals_excel(self, token, fpath):
        if not os.path.isfile(fpath):
            book = openpyxl.Workbook()
            sheet = book.active
            sheet.append(
                ["Skipped QRN"])
        else:
            book = openpyxl.load_workbook(fpath)
            sheet = book.active

        sheet.append(token)
        book.save(fpath)

    def create_excel(self, record, fpath):
        if not os.path.isfile(fpath):
            book = openpyxl.Workbook()
            sheet = book.active
            sheet.append(
                ["FinnOne No", "Customer Name", "State", "Branch", "Disbursement Date", "Maker ID", "Maker date",
                 "Checker ID", "Checker Date", "Final Checker date", "STATUS (Successful/Rejected)", "Remarks", "QRN",
                 "Security Interest ID", "Asset ID", "TAT (Maker date-Checker date)",
                 "TAT (Checker date-Final checker date)", "Message"])
        else:
            book = openpyxl.load_workbook(fpath)
            sheet = book.active

        sheet.append(record)
        book.save(fpath)

    def get_date(self):
        dt = datetime.datetime.now()
        date = dt.strftime("%d-%m-%Y")
        return str(date)


def make_dir(*paths):
    # Creates all required directories if not present, mentioned in the path.
    for pt in paths:
        if not (os.path.isdir(pt)):
            try:
                os.makedirs(pt, mode=0o777, exist_ok=True)
            except OSError as exception:
                if exception.errno != errno.EEXIST:
                    raise


if __name__ == "__main__":
    SUCCESS_PATH = os.path.join(os.getcwd(), "Files", "Excel", "Successful")
    FAIL_PATH = os.path.join(os.getcwd(), "Files", "Excel", "Failed")
    TOKEN_PATH = os.path.join(os.getcwd(), "Files", "Excel", "MIS to be Processed")
    REJECT_PATH = os.path.join(os.getcwd(), "Files", "Excel", "Rejected")
    IGNORE_PATH = os.path.join(os.getcwd(), "Files", "Excel", "Ignore token")
    Driver_path = os.path.join(os.getcwd(), "Files", "Chrome_driver")
    util_path = os.path.join(os.getcwd(), "Files", "util")
    make_dir(Driver_path, util_path, SUCCESS_PATH, FAIL_PATH, TOKEN_PATH,REJECT_PATH,IGNORE_PATH)
    IGNORE_FILE = os.path.join(IGNORE_PATH,"ignored_tokens.xlsx")
    cpath = os.path.join(Driver_path, "chromedriver.exe")
    app = wx.App()
    frame = Checker_Process(parent=None, id=-1)
    frame.Show()
    app.MainLoop()
